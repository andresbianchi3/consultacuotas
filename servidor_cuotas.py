from flask import Flask, render_template_string, request, jsonify
import openpyxl
import json
import os
import re
from datetime import datetime

app = Flask(__name__)
JSON_FILE = 'cuotas.json'

# HTML de la interfaz para cargar JSON
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cargar JSON de Cuotas</title>
    <style>
        * { box-sizing: border-box; }
        body {
            margin: 0;
            font-family: Arial, sans-serif;
            background: linear-gradient(135deg, #e0f2fe, #f8fafc);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 24px;
        }
        .container {
            max-width: 620px;
            width: 100%;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 50px rgba(15, 23, 42, 0.12);
            padding: 32px;
        }
        h1 {
            margin: 0 0 8px;
            color: #0f172a;
            font-size: 30px;
        }
        .subtitle {
            margin: 0 0 24px;
            color: #475569;
            font-size: 14px;
        }
        .info {
            background: #eff6ff;
            border-left: 4px solid #2563eb;
            color: #1d4ed8;
            border-radius: 8px;
            padding: 12px 14px;
            margin-bottom: 22px;
            font-size: 13px;
        }
        .upload-area {
            border: 2px dashed #2563eb;
            border-radius: 12px;
            padding: 38px 18px;
            text-align: center;
            background: #f8fbff;
            cursor: pointer;
            transition: background 0.2s ease, transform 0.2s ease;
        }
        .upload-area:hover { background: #eef5ff; }
        .upload-area.dragover {
            background: #e8f0ff;
            transform: scale(1.01);
        }
        .upload-icon { font-size: 42px; margin-bottom: 8px; }
        .upload-text { font-weight: 700; color: #0f172a; margin-bottom: 6px; }
        .upload-hint { color: #64748b; font-size: 13px; }
        input[type="file"] { display: none; }
        .status {
            display: none;
            align-items: center;
            gap: 12px;
            border-radius: 10px;
            padding: 14px 16px;
            margin-top: 22px;
        }
        .status.success {
            display: flex;
            background: #dcfce7;
            border: 1px solid #86efac;
            color: #166534;
        }
        .status.error {
            display: flex;
            background: #fee2e2;
            border: 1px solid #fca5a5;
            color: #991b1b;
        }
        .status-icon {
            font-size: 22px;
        }
        .status-text {
            font-size: 14px;
            line-height: 1.4;
        }
        .rules {
            margin-top: 16px;
            font-size: 12px;
            color: #475569;
            line-height: 1.5;
        }
        .rules strong { color: #0f172a; }
    </style>
</head>
<body>
    <div class="container">
        <h1>� Cargar datos de cuotas</h1>
        <p class="subtitle">Elegí qué tipo de archivo querés cargar y actualizá la base de cuotas.</p>

        <div class="tabs" style="display:flex; gap:12px; margin-bottom:20px; flex-wrap:wrap;">
            <button class="tab-btn active" data-mode="json" type="button">JSON</button>
            <button class="tab-btn" data-mode="excel" type="button">Excel</button>
        </div>

        <div class="info">
            El archivo se guardará en <strong>cuotas.json</strong> dentro de esta carpeta.
        </div>

        <div class="panel-upload active" data-panel="json">
            <div class="upload-area" id="uploadAreaJson">
                <div class="upload-icon">📁</div>
                <div class="upload-text">Arrastra el JSON aquí</div>
                <div class="upload-hint">o hacé clic para seleccionar un archivo .json</div>
                <input type="file" id="fileInputJson" accept=".json,application/json">
            </div>
        </div>

        <div class="panel-upload" data-panel="excel" style="display:none;">
            <div class="upload-area" id="uploadAreaExcel">
                <div class="upload-icon">📊</div>
                <div class="upload-text">Arrastra el Excel aquí</div>
                <div class="upload-hint">o hacé clic para seleccionar un archivo .xlsx o .xls</div>
                <input type="file" id="fileInputExcel" accept=".xlsx,.xls">
            </div>
        </div>

        <div class="status success" id="statusSuccess">
            <div class="status-icon">✓</div>
            <div class="status-text" id="statusMsg"></div>
        </div>

        <div class="status error" id="statusError">
            <div class="status-icon">⚠️</div>
            <div class="status-text" id="errorMsg"></div>
        </div>

        <div class="rules" id="rulesJson">
            <strong>Formato JSON esperado:</strong> un array de afiliados o un objeto con la clave <strong>afiliados</strong>.
        </div>

        <div class="rules" id="rulesExcel" style="display:none;">
            <strong>Formato Excel esperado:</strong> hojas <strong>MANADA</strong>, <strong>UNIDAD</strong>, <strong>CAMINANTES</strong> y <strong>ROVERS</strong> con la columna <strong>Nombre</strong>.
        </div>
    </div>

    <script>
        const tabButtons = document.querySelectorAll('.tab-btn');
        const panels = document.querySelectorAll('.panel-upload');
        const statusSuccess = document.getElementById('statusSuccess');
        const statusError = document.getElementById('statusError');
        const statusMsg = document.getElementById('statusMsg');
        const errorMsg = document.getElementById('errorMsg');

        let activeMode = 'json';

        function setMode(mode) {
            activeMode = mode;
            tabButtons.forEach(btn => btn.classList.toggle('active', btn.dataset.mode === mode));
            panels.forEach(panel => {
                const isActive = panel.dataset.panel === mode;
                panel.style.display = isActive ? 'block' : 'none';
                panel.classList.toggle('active', isActive);
            });
            document.getElementById('rulesJson').style.display = mode === 'json' ? 'block' : 'none';
            document.getElementById('rulesExcel').style.display = mode === 'excel' ? 'block' : 'none';
            statusSuccess.style.display = 'none';
            statusError.style.display = 'none';
        }

        tabButtons.forEach(btn => {
            btn.addEventListener('click', () => setMode(btn.dataset.mode));
        });

        function bindUpload(uploadArea, fileInput, mode) {
            uploadArea.addEventListener('click', () => fileInput.click());
            uploadArea.addEventListener('dragover', (e) => {
                e.preventDefault();
                uploadArea.classList.add('dragover');
            });
            uploadArea.addEventListener('dragleave', () => {
                uploadArea.classList.remove('dragover');
            });
            uploadArea.addEventListener('drop', (e) => {
                e.preventDefault();
                uploadArea.classList.remove('dragover');
                if (e.dataTransfer.files[0]) {
                    fileInput.files = e.dataTransfer.files;
                    procesarArchivo(mode, fileInput.files[0]);
                }
            });
            fileInput.addEventListener('change', () => procesarArchivo(mode, fileInput.files[0]));
        }

        const uploadAreaJson = document.getElementById('uploadAreaJson');
        const fileInputJson = document.getElementById('fileInputJson');
        const uploadAreaExcel = document.getElementById('uploadAreaExcel');
        const fileInputExcel = document.getElementById('fileInputExcel');

        bindUpload(uploadAreaJson, fileInputJson, 'json');
        bindUpload(uploadAreaExcel, fileInputExcel, 'excel');

        function procesarArchivo(mode, file) {
            if (!file) return;

            const formData = new FormData();
            formData.append('file', file);

            statusSuccess.style.display = 'none';
            statusError.style.display = 'none';
            statusMsg.textContent = 'Procesando...';
            statusSuccess.style.display = 'flex';

            fetch('/procesar', {
                method: 'POST',
                body: formData
            })
            .then(r => r.json())
            .then(data => {
                if (data.success) {
                    statusSuccess.style.display = 'flex';
                    statusMsg.textContent = data.message;
                    statusError.style.display = 'none';
                } else {
                    statusSuccess.style.display = 'none';
                    statusError.style.display = 'flex';
                    errorMsg.textContent = data.error;
                }
            })
            .catch(err => {
                statusSuccess.style.display = 'none';
                statusError.style.display = 'flex';
                errorMsg.textContent = err.message;
            });
        }

        setMode('json');
    </script>
</body>
</html>
'''

def formato_pesos(valor):
    """Formatea número a pesos argentinos: $ 28.000,00"""
    if valor == 0:
        return '$ 0.00'
    return f'$ {valor:,}.00'.replace(',', '.')


def parsear_numero(valor):
    """Convierte valores numéricos o textos de Excel a entero sin romper con fórmulas."""
    if valor is None or valor == '':
        return 0

    if isinstance(valor, bool):
        return 0

    if isinstance(valor, (int, float)):
        return int(float(valor))

    if isinstance(valor, str):
        texto = valor.strip()
        if texto == '':
            return 0
        if texto.startswith('='):
            return 0

        texto = texto.replace('$', '').replace(' ', '')
        if texto.startswith('(') and texto.endswith(')'):
            texto = '-' + texto[1:-1]

        # Elimina símbolo de moneda y otras letras no numéricas
        texto = re.sub(r'[^0-9,.-]', '', texto)
        if texto in ('', '-', '.', '-.', ',', '-,'):
            return 0

        texto = texto.replace('.', '').replace(',', '.')
        try:
            return int(float(texto))
        except ValueError:
            return 0

    return 0


def procesar_excel(filepath):
    """Procesa el Excel y retorna la lista de datos"""
    # Cargar con data_only=True para obtener valores en lugar de fórmulas
    wb = openpyxl.load_workbook(filepath, data_only=True)
    datos = []
    ramas = ['MANADA', 'UNIDAD', 'CAMINANTES', 'ROVERS']
    
    for rama in ramas:
        if rama not in wb.sheetnames:
            continue
            
        ws = wb[rama]
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 2:
            continue
        
        # Fila 1 (índice 1) tiene los headers
        headers = rows[1]
        
        # Encontrar la columna de Nombre
        try:
            nombre_col = headers.index('Nombre')
        except ValueError:
            continue
        
        # Procesar desde la fila 2 en adelante
        for row in rows[2:]:
            if not row or not row[nombre_col]:
                break
            
            nombre = str(row[nombre_col]).strip()
            if not nombre:
                break
            
            # Obtener valores de meses (columnas 3-7: Abril, Mayo, Junio, Julio, Agosto)
            abril = parsear_numero(row[3] if len(row) > 3 else None)
            mayo = parsear_numero(row[4] if len(row) > 4 else None)
            junio = parsear_numero(row[5] if len(row) > 5 else None)
            julio = parsear_numero(row[6] if len(row) > 6 else None)
            agosto = parsear_numero(row[7] if len(row) > 7 else None)
            
            persona = {
                'Nombre': nombre,
                'Rama': rama,
                'Abril': formato_pesos(abril),
                'Mayo': formato_pesos(mayo),
                'Junio': formato_pesos(junio),
                'Julio': formato_pesos(julio),
                'Agosto': formato_pesos(agosto)
            }
            
            datos.append(persona)
    
    return datos

@app.route('/')
@app.route('/cargarcuotasexcel')
@app.route('/cargarcuotasexcel/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/procesar', methods=['POST'])
def procesar():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió archivo'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Archivo vacío'})

        filename = file.filename.lower()
        temp_path = 'temp_upload'

        if filename.endswith('.json'):
            file.save(temp_path + '.json')
            with open(temp_path + '.json', 'r', encoding='utf-8') as f:
                datos = json.load(f)

            if isinstance(datos, dict) and 'afiliados' in datos:
                payload = datos['afiliados']
            elif isinstance(datos, list):
                payload = datos
            else:
                return jsonify({'success': False, 'error': 'El JSON no tiene el formato esperado. Debe ser una lista o un objeto con "afiliados".'})

            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=4)

            os.remove(temp_path + '.json')
            return jsonify({
                'success': True,
                'message': f'JSON cargado correctamente. Se guardaron {len(payload) if isinstance(payload, list) else 1} registros en cuotas.json.'
            })

        file.save(temp_path + '.xlsx')
        datos = procesar_excel(temp_path + '.xlsx')
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=4)

        os.remove(temp_path + '.xlsx')
        preview = datos[:8] if datos else []

        return jsonify({
            'success': True,
            'message': f'{len(datos)} personas procesadas y guardadas en cuotas.json',
            'preview': preview
        })

    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'El archivo JSON no es válido.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/descargar')
def descargar():
    if os.path.exists(JSON_FILE):
        return open(JSON_FILE, 'rb'), 200, {'Content-Disposition': f'attachment; filename=cuotas.json'}
    return jsonify({'error': 'Archivo no encontrado'}), 404

if __name__ == '__main__':
    print('=' * 60)
    print('🚀 SERVIDOR DE CUOTAS INICIADO')
    print('=' * 60)
    print()
    print('📍 Abre en tu navegador:')
    print('   http://localhost:5000/cargarcuotasexcel')
    print()
    print('✅ El archivo se guardará en: cuotas.json')
    print('⚠️  Asegúrate de estar en la carpeta donde guardas tu JSON')
    print()
    print('Presiona CTRL+C para detener el servidor')
    print('=' * 60)
    print()
    
    app.run(debug=False, host='localhost', port=5000)